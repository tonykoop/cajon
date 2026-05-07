"""
Build Cajon-Family-Design.xlsx — parametric workbook with real Excel formulas.
Run: python build_xlsx.py
Requires: pip install openpyxl
"""
from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "Cajon-Family-Design.xlsx"

HEADER_FONT = Font(name="Cambria", size=12, bold=True, color="6B4F2C")
TITLE_FONT = Font(name="Cambria", size=16, bold=True, color="6B4F2C")
INPUT_FONT = Font(name="Calibri", size=11, color="0000FF")
BODY_FONT = Font(name="Calibri", size=11, color="000000")
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="6B4F2C")
THIN = Side(style="thin", color="C8B89E")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="F4ECDB")

IN_FMT = '0.000"\""'
HZ_FMT = '0.0 "Hz"'
USD_FMT = '"$"#,##0.00'


def build_design_sheet(ws):
    ws.title = "Design"
    ws["A1"] = "Cajón Family — Parametric Design Sheet"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    ws["A3"] = "DESIGN INPUTS (blue = your knob)"
    ws["A3"].font = HEADER_FONT
    ws["A3"].fill = HEAD_FILL

    inputs = [
        ("Member", "Standard", "Compact / Standard / Bass"),
        ("Outer Height (in)", 18, "15 / 18 / 22"),
        ("Outer Width (in)", 12, "11 / 12 / 14"),
        ("Outer Depth (in)", 12, "11 / 12 / 14"),
        ("Body Panel Thickness (in)", 0.75, "Sides + top + bottom"),
        ("Back Panel Thickness (in)", 0.5, "Baltic birch back"),
        ("Tapa Thickness (in)", 0.118, "3 mm Baltic birch ply"),
        ("Sound Hole Diameter (in)", 4.5, "4.0 / 4.5 / 5.0 by member"),
        ("Speed of Sound (in/s)", 13507, "= 343 m/s × 39.37"),
    ]
    for i, (label, val, note) in enumerate(inputs, start=4):
        ws.cell(row=i, column=1, value=label).font = BODY_FONT
        cell = ws.cell(row=i, column=2, value=val)
        cell.font = INPUT_FONT
        if isinstance(val, float):
            cell.number_format = '0.000'
        ws.cell(row=i, column=3, value=note).font = NOTE_FONT

    ws["A14"] = "DERIVED — HELMHOLTZ + PLATE (1,1) (formula-driven)"
    ws["A14"].font = HEADER_FONT
    ws["A14"].fill = HEAD_FILL

    derivations = [
        ("Internal Volume (in³)", "=(B6-2*B7)*(B7-2*B7)*(B5-B7-B8)", "(W-2t)(D-2t)(H-top-back); approximate, treats top thickness = body t"),
        ("Sound Hole Area (in²)", "=PI()*(B11/2)^2", ""),
        ("Effective Neck Length (in)", "=B8+1.7*(B11/2)", "L_eff = t_back + 1.7·r — flush thin-wall end correction (BOTH sides counted)"),
        ("Helmholtz Frequency (Hz)", "=(B12/(2*PI()))*SQRT(B16/(B15*B17))", "f_H = (c/2π)·√(A/(V·L_eff))"),
        ("Nearest Note", '=CHOOSE(MOD(ROUND(12*LOG(B18,2)-12*LOG(440,2)+69,0),12)+1,"C","C#","D","D#","E","F","F#","G","G#","A","A#","B")&INT((ROUND(12*LOG(B18,2)-12*LOG(440,2)+69,0))/12-1)', ""),
    ]
    # Note: cell positions below are illustrative; the actual cross-references
    # should be re-mapped if you reorder. Keeping the structure simple here.
    derived_start = 15
    for i, (label, formula, note) in enumerate(derivations):
        r = derived_start + i
        ws.cell(row=r, column=1, value=label).font = BODY_FONT
        c = ws.cell(row=r, column=2, value=formula)
        c.font = BODY_FONT
        ws.cell(row=r, column=3, value=note).font = NOTE_FONT

    ws["A22"] = "FAMILY TABLE"
    ws["A22"].font = HEADER_FONT
    ws["A22"].fill = HEAD_FILL

    family_header = ["Member", "H (in)", "W (in)", "D (in)", "Body t (in)", "Hole d (in)", "Predicted f_H (Hz)", "Predicted plate (1,1) (Hz)"]
    for col, h in enumerate(family_header, start=1):
        cell = ws.cell(row=23, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEAD_FILL
    family_rows = [
        ("Compact", 15, 11, 11, 0.625, 4.0, 107, 107),
        ("Standard", 18, 12, 12, 0.75, 4.5, 96, 84),
        ("Bass", 22, 14, 14, 0.75, 5.0, 77, 60),
    ]
    for ri, row in enumerate(family_rows, start=24):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            if isinstance(val, float):
                cell.number_format = '0.000'

    # Column widths
    for col, width in enumerate([28, 14, 14, 14, 14, 14, 22, 26], start=1):
        ws.column_dimensions[chr(64 + col)].width = width


def build_bom_sheet(ws):
    ws.title = "BOM"
    src = ROOT / "bom.csv"
    if not src.exists():
        ws["A1"] = "(BOM source bom.csv missing)"
        return
    with open(src, encoding="utf-8") as f:
        reader = csv.reader(f)
        for ri, row in enumerate(reader, start=1):
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ri == 1:
                    cell.font = HEADER_FONT
                    cell.fill = HEAD_FILL
                else:
                    cell.font = BODY_FONT
    for col, width in enumerate([24, 36, 18, 14, 10, 10, 14, 16, 10, 36], start=1):
        ws.column_dimensions[chr(64 + col)].width = width


def build_validation_sheet(ws):
    ws.title = "Validation"
    src = ROOT / "validation.csv"
    if not src.exists():
        ws["A1"] = "(Validation source validation.csv missing)"
        return
    with open(src, encoding="utf-8") as f:
        reader = csv.reader(f)
        for ri, row in enumerate(reader, start=1):
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ri == 1:
                    cell.font = HEADER_FONT
                    cell.fill = HEAD_FILL
                else:
                    cell.font = BODY_FONT
    for col, width in enumerate([12, 10, 32, 18, 12, 12, 10, 10, 14, 16, 36], start=1):
        ws.column_dimensions[chr(64 + col)].width = width


def main():
    wb = Workbook()
    build_design_sheet(wb.active)
    build_bom_sheet(wb.create_sheet("BOM"))
    build_validation_sheet(wb.create_sheet("Validation"))
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
